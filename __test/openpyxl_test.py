import os
import sys
import pprint

import openpyxl

def readStockXL(_file):
    try:
        if os.path.exists(_file) == False:
            print('The file %s is not exit!!!'%(_file))
            return None
        wb = openpyxl.load_workbook(filename=_file, read_only=True)
        #sheetList = wb.get_sheet_names()
        ws = wb.active
        #nrows = ws.get_highest_row()
        #ncols = ws.get_highest_column()

        _rowsData = list(ws.rows)
        _colsData = list(ws.columns)
        nrows = len(_rowsData)
        ncols = len(_colsData)
        print('ROW: %d ,COL: %d'%(nrows, ncols))
        colnames = list(_rowsData[0])

        data=[]
        for row in _rowsData[1:]:
            tmp = {}
            for cun in range(ncols):
                tmp[colnames[cun].value] = row[cun].value
            data.append(tmp)

        return data
    except Exception as err:
        print(">>>>>> Exception: " + str(err))
        return None
    finally:
        None

def writeStockXL(_file, _data):
    try:
        if _data == None:
            return None

        wb = openpyxl.Workbook()
        #ws = wb.active
        ws = wb.create_sheet(title='mySheet1',index = 0) #creat a new sheet and insert at first position

        stockList = sorted(_data.keys())
        colNames = sorted(_data[stockList[0]].keys())
        #colNames.remove('name')
        ws.append(colNames)

        stockData = []
        for i in stockList:
            data = _data[i]
            rowDatas = []
            for j in colNames:
                rowDatas.append(data[j])
            stockData.append(rowDatas)
        
        for row in stockData:
           ws.append(row)

        wb.save(_file)
        return 0
    except Exception as err:
        print(">>>>>> Exception: " + str(err))
        return -1
    finally:
        None

def writeXL(_file):
    try:
        wb = openpyxl.Workbook()
        #ws = wb.active
        ws = wb.create_sheet(title='mySheet1',index = 0) #creat a new sheet and insert at first position
        
        rows = [
            ['Number', 'Batch 1', 'Batch 2'],
            [1, 40, 300],
            [2, 40, 300],
            [3, 40, 250],
            [4, 50, 300],
            [5, 30, 100],
            [6, 25, 500],
            [7, 50, 100],
        ]
        
        for row in rows:
           ws.append(row)

        '''
        add img to sheet
        img = openpyxl.drawing.image.Image('0001.jpg')
        ws.add_image(img, 'A9')
        '''
        '''
        add chart to sheet
        chart = openpyxl.chart.AreaChart()
        chart.title = "Area Chart"
        chart.style = 13
        chart.x_axis.title = 'Test'
        chart.y_axis.title = 'Percentage'
        cats = openpyxl.chart.Reference(ws, min_col=1, min_row=1, max_row=7)
        data = openpyxl.chart.Reference(ws, min_col=1, min_row=1, max_col=3, max_row=7)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        ws.add_chart(chart, 'A10')
        '''

        wb.save(_file)
        return 0
    except Exception as err:
        print(">>>>>> Exception: " + str(err))
        return -1
    finally:
        None

if __name__ == '__main__':
    writeXL('fdf.xlsx')
    pprint.pprint(readStockXL('fdf.xlsx'))
